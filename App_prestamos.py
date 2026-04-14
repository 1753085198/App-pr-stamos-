import streamlit as st
from streamlit_gsheets import GSheetsConnection
import pandas as pd
import uuid
import io
import time
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime
from dateutil.relativedelta import relativedelta
import requests
import base64
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# 1. CONFIGURACIÓN INICIAL
st.set_page_config(page_title="SISTEMA FINANCIERO COOP SAN BLAS", page_icon="🏦", layout="wide")

# 2. CSS PARA INTERFAZ GIGANTE
st.markdown("""
    <style>
    .stMarkdown p, label, .stNumberInput label, .stTextInput label { font-size: 26px !important; font-weight: 700 !important; }
    input { font-size: 22px !important; height: 50px !important; }
    .stDownloadButton>button { font-size: 28px !important; font-weight: 800 !important; height: 5rem !important; border-radius: 15px !important; background-color: #1D6F42 !important; color: white !important; }
    .stButton>button { font-size: 24px !important; font-weight: 700 !important; border-radius: 12px !important; }
    .stButton>button[kind="secondary"] { background-color: #dc3545 !important; color: white !important; }
    div.stButton > button:first-child[key^="btn_nuevo"] { background-color: #ff5722 !important; color: white !important; border-radius: 50px !important; padding: 20px 40px !important; font-size: 28px !important; font-weight: 900 !important; position: fixed; bottom: 30px; right: 30px; z-index: 9999; border: 3px solid white !important; }
    [data-testid="stMetricValue"] { font-size: 65px !important; font-weight: 900 !important; color: #007bff !important; }
    </style>
""", unsafe_allow_html=True)

conn = st.connection("gsheets", type=GSheetsConnection)

# --- ESTADOS GLOBALES ---
if 'pago_key' not in st.session_state: st.session_state.pago_key = 0
if 'id_abierto' not in st.session_state: st.session_state.id_abierto = None

# --- FUNCIONES CORE ---
def cargar(h):
    try:
        df = conn.read(worksheet=h, ttl=0)
        if df is not None and "ID" in df.columns:
            df["ID"] = df["ID"].astype(str).str.replace(".0", "", regex=False)
        return df if df is not None else pd.DataFrame()
    except: return pd.DataFrame()

def subir_img(archivo):
    try:
        res = requests.post("https://api.imgbb.com/1/upload", data={"key": st.secrets["IMGBB_API_KEY"], "image": base64.b64encode(archivo).decode('utf-8')})
        return res.json()["data"]["url"]
    except: return ""

def generar_excel_grupal(df, titulo):
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine='openpyxl') as writer:
        ws = writer.book.create_sheet("REPORTE_GENERAL", 0)
        header_fill = PatternFill(start_color="0B2F45", end_color="0B2F45", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        v_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid") 
        r_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") 
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        label_monto = "DEUDA PENDIENTE" if "PRÉSTAMOS" in titulo.upper() else "TOTAL APORTADO"
        ws.merge_cells("A1:D1"); ws["A1"] = f"COOP SAN BLAS - REPORTE DE {titulo}"; ws["A1"].font = Font(bold=True, size=16, color="0B2F45"); ws["A1"].alignment = center_align
        ws.append([])

        headers = ["NOMBRE COMPLETO", "CÉDULA", label_monto, "ESTADO ACTUAL"]
        ws.append(headers)
        for cell in ws[3]: cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align; cell.border = thin_border

        for _, row in df.iterrows():
            m = float(row.get('Saldo_Total_Aportado', row.get('Saldo_Restante', 0)))
            if "PRÉSTAMOS" in titulo.upper():
                estado = row.get("Estado", "ACTIVO"); color = v_fill if estado == "PAGADO" else r_fill
            else:
                estado = "AL DÍA" if m > 0 else "PENDIENTE"; color = v_fill if m > 0 else r_fill
            ws.append([row['Nombre'], row['Cedula'], m, estado])
            for cell in ws[ws.max_row]: cell.fill = color; cell.alignment = center_align; cell.border = thin_border

        ws.column_dimensions['A'].width = 40; ws.column_dimensions['B'].width = 20; ws.column_dimensions['C'].width = 25; ws.column_dimensions['D'].width = 25
    return out.getvalue()

def generar_excel_personal(row, historial, tipo):
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine='openpyxl') as writer:
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        header_font = Font(bold=True)
        right_align = Alignment(horizontal="right")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        p_fil = pd.DataFrame()
        if not historial.empty:
            col_id = 'ID_Socio' if 'ID_Socio' in historial.columns else 'ID_Prestamo'
            if col_id in historial.columns: p_fil = historial[historial[col_id] == row['ID']]

        if tipo == "PRÉSTAMO":
            ws1 = writer.book.create_sheet("AMORTIZACION_SAN_BLAS", 0)
            
            # --- BLINDAJE DE FECHA DE INICIO ---
            fecha_val = row.get('Fecha_Inicio')
            if pd.isna(fecha_val) or str(fecha_val).strip() in ["", "nan", "NaT", "None"]:
                f_inicio = datetime.now()
            else:
                try:
                    f_inicio = pd.to_datetime(fecha_val).to_pydatetime()
                except:
                    f_inicio = datetime.now()
            # -----------------------------------
            
            meses_plazo = int(row.get('Meses_Totales', 1))
            f_vencimiento = f_inicio + relativedelta(months=meses_plazo)
            
            hoy = datetime.now()
            diff = relativedelta(hoy, f_inicio)
            meses_transcurridos = diff.years * 12 + diff.months
            meses_a_pagar = min(meses_transcurridos, meses_plazo)
            
            pagos_reales = int(row.get('Pagos_Realizados', 0))
            cuota_val = float(row.get('Cuota_Mensual', 0))
            debe_estar_pagado = meses_a_pagar * cuota_val
            lo_que_ha_pagado = pagos_reales * cuota_val
            faltante_dia = max(0, debe_estar_pagado - lo_que_ha_pagado)

            info = [
                ("COOP SAN BLAS - ESTADO DE AYUDA", ""),
                ("NOMBRE", row['Nombre']),
                ("CEDULA", row['Cedula']),
                ("FECHA INICIAL", f_inicio.strftime("%d/%m/%Y")),
                ("FECHA VENCIMIENTO", f_vencimiento.strftime("%d/%m/%Y")),
                ("PLAZO MESES", meses_plazo),
                ("CONTRIBUCION ANUAL", "8%"),
                ("VALOR AYUDA REEMBOLSABLE", round(float(row.get('Monto_Inicial', 0)), 2)),
                ("SUMA TOTAL CON INTERÉS", round(cuota_val * meses_plazo, 2)),
                ("--- ESTADO DE CUENTA A LA FECHA ---", ""),
                ("MESES TRANSCURRIDOS", meses_transcurridos),
                ("CUOTAS QUE DEBERÍA TENER", meses_a_pagar),
                ("CUOTAS PAGADAS REALES", pagos_reales),
                ("VALOR PARA ESTAR AL DÍA", f"$ {round(faltante_dia, 2)}")
            ]
            
            for idx, (label, val) in enumerate(info, start=1):
                ws1.cell(row=idx, column=2, value=label).font = header_font
                cell_val = ws1.cell(row=idx, column=3, value=val)
                if label == "CONTRIBUCION ANUAL": cell_val.alignment = right_align
                ws1.cell(row=idx, column=2).border = thin_border
                ws1.cell(row=idx, column=3).border = thin_border

            ws1.append([])
            headers = ["N°", "FECHAS", "SALDO", "CAPITAL", "CONTRIB", "CUOTA", "PAGADO", "DIFERENCIA"]
            ws1.append(headers)
            for cell in ws1[ws1.max_row]: 
                cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align; cell.border = thin_border

            monto_ini = float(row.get('Monto_Inicial', 0))
            cap_m = monto_ini / meses_plazo if meses_plazo > 0 else 0
            int_m = (cuota_val * meses_plazo - monto_ini) / meses_plazo if meses_plazo > 0 else 0
            
            for i in range(1, meses_plazo + 1):
                ws1.append([i, "", "---", round(cap_m, 2), round(int_m, 2), round(cuota_val, 2), 
                            (round(cuota_val, 2) if i <= pagos_reales else 0), "---"])
                for cell in ws1[ws1.max_row]: cell.border = thin_border; cell.alignment = center_align

            for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']: ws1.column_dimensions[col].width = 22

        else:
            # COOPERATIVA Y AYUDAS
            ws1 = writer.book.create_sheet("RESUMEN_FINANCIERO", 0)
            ws1.merge_cells("A1:B1"); ws1["A1"] = f"COOP SAN BLAS - {tipo}"; ws1["A1"].font = Font(bold=True, size=16, color="0B2F45"); ws1["A1"].alignment = center_align
            ws1.append([f"SOCIO:", row['Nombre']]); ws1.append([f"CÉDULA:", row['Cedula']]); ws1.append([""])

            ws1.append(["FECHA DEL APORTE", "MONTO APORTADO ($)"])
            for cell in ws1[5]: cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align; cell.border = thin_border

            for _, p in p_fil.iterrows(): 
                f = p.get('Fecha', ''); m = p.get('Monto', 0)
                ws1.append([f, m])
                for cell in ws1[ws1.max_row]: cell.border = thin_border; cell.alignment = center_align
            
            ws1.append(["TOTAL ACUMULADO:", row.get('Saldo_Total_Aportado', 0)])
            ws1.cell(row=ws1.max_row, column=1).font = Font(bold=True); ws1.cell(row=ws1.max_row, column=2).font = Font(bold=True)
            ws1.column_dimensions['A'].width = 25; ws1.column_dimensions['B'].width = 25

            ws2 = writer.book.create_sheet("RESPALDOS_DIGITALES", 1)
            ws2.merge_cells("A1:B1"); ws2["A1"] = "COMPROBANTES DIGITALES"; ws2["A1"].font = Font(bold=True, size=16, color="0B2F45"); ws2["A1"].alignment = center_align
            ws2.append([""]); ws2.append(["FECHA DE PAGO", "ENLACE AL RECIBO (CLIC PARA ABRIR)"])
            for cell in ws2[3]: cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align; cell.border = thin_border
            
            for _, p in p_fil.iterrows(): 
                f = p.get('Fecha', ''); c = p.get('Comprobante', 'N/A')
                ws2.append([f, c])
                for cell in ws2[ws2.max_row]: 
                    cell.border = thin_border; cell.alignment = center_align
                    if str(cell.value).startswith('http'):
                        cell.font = Font(color="0563C1", underline="single"); cell.hyperlink = cell.value

            ws2.column_dimensions['A'].width = 25; ws2.column_dimensions['B'].width = 60
            
    return out.getvalue()

def enviar_mail(dest, nom, exc, url, tipo):
    try:
        msg = MIMEMultipart(); msg['From'] = st.secrets["EMAIL_USER"]; msg['To'] = dest; msg['Subject'] = f"✅ Reporte Financiero Coop San Blas - {tipo}"
        msg.attach(MIMEText(f"Estimado(a) {nom},\n\nSe adjunta su estado de cuenta actualizado de Coop San Blas.\n\nPuede visualizar su recibo aquí: {url}\n\nAtentamente,\nAdministración Coop San Blas.", 'plain'))
        p = MIMEBase('application', 'octet-stream'); p.set_payload(exc); encoders.encode_base64(p)
        p.add_header('Content-Disposition', f"attachment; filename=Reporte_CoopSanBlas_{tipo}.xlsx"); msg.attach(p)
        s = smtplib.SMTP('smtp.gmail.com', 587); s.starttls(); s.login(st.secrets["EMAIL_USER"], st.secrets["EMAIL_PASS"]); s.send_message(msg); s.quit()
    except: pass

# --- NAVEGACIÓN ---
with st.sidebar:
    st.markdown("# 🏦 COOP SAN BLAS")
    sec = st.radio("SECCIONES:", ["💰 PRÉSTAMOS", "🤝 COOPERATIVA", "🚑 AYUDAS ECON."], index=0)
    
    st.write("---")
    st.write("### 📂 IMPORTAR DATOS")
    archivo_antiguo = st.file_uploader("Sube Excel antiguo aquí", type=["xlsx", "csv"])
    if archivo_antiguo:
        st.info("Archivo cargado. Copia estos datos a tu Google Sheets (columna Fecha_Inicio) para integrarlos de forma permanente.")

# --- MODO PRÉSTAMOS ---
if sec == "💰 PRÉSTAMOS":
    st.title("💰 GESTIÓN DE PRÉSTAMOS - SAN BLAS")
    df_p, df_h = cargar("Prestamos"), cargar("Pagos")
    if not df_p.empty: st.download_button("📊 EXCEL GENERAL", data=generar_excel_grupal(df_p, "PRÉSTAMOS"), file_name="Reporte_SanBlas_Prestamos.xlsx", use_container_width=True)
    
    if st.button("👤 NUEVO PRÉSTAMO", key="btn_nuevo_p"): st.session_state.show_form_p = not st.session_state.get('show_form_p', False)
    if st.session_state.get('show_form_p'):
        with st.form("form_p"):
            n, c, e = st.text_input("Nombre:"), st.text_input("Cédula:"), st.text_input("Email:")
            m, t, p = st.number_input("Monto:", min_value=1.0), st.number_input("Tasa Anual %:", value=8.0), st.number_input("Meses:", value=36)
            f_ini = st.date_input("Fecha de Inicio:", value=datetime.today())
            if st.form_submit_button("💾 GUARDAR"):
                interes_total = m * (t / 100) * (p / 12)
                deuda_total = m + interes_total
                cuo = deuda_total / p
                new = pd.DataFrame([{"ID":str(uuid.uuid4())[:8], "Nombre":n, "Cedula":c, "Email":e, "Monto_Inicial":m, "Saldo_Restante":round(deuda_total,2), "Cuota_Mensual":round(cuo,2), "Meses_Totales":p, "Pagos_Realizados":0, "Estado":"ACTIVO", "Fecha_Inicio": f_ini.strftime("%Y-%m-%d")}])
                conn.update(worksheet="Prestamos", data=pd.concat([df_p, new], ignore_index=True)); st.session_state.show_form_p = False; st.rerun()

    bq_p = st.text_input("🔍 BUSCAR PRESTAMISTA:")
    act_p = df_p[df_p["Estado"]=="ACTIVO"] if not df_p.empty else pd.DataFrame()
    if bq_p and not act_p.empty: act_p = act_p[act_p['Nombre'].str.contains(bq_p, case=False)]

    for idx, row in act_p.iterrows():
        with st.expander(f"👤 {row['Nombre'].upper()} | SALDO: ${row['Saldo_Restante']}", expanded=(st.session_state.id_abierto == row['ID'])):
            c1, c2 = st.columns(2)
            with c1: 
                st.write(f"**CUOTA:** ${row['Cuota_Mensual']} | **PROGRESO:** {row['Pagos_Realizados']}/{row['Meses_Totales']}")
                with st.form(key=f"fp_{row['ID']}"):
                    ft = st.file_uploader("📸 RECIBO:", key=f"ip_{row['ID']}_{st.session_state.pago_key}")
                    if st.form_submit_button("✅ REGISTRAR PAGO"):
                        if ft:
                            url = subir_img(ft.getvalue())
                            st.session_state.id_abierto = row['ID']
                            new_h = pd.DataFrame([{"ID_Prestamo": row['ID'], "Fecha": datetime.now().strftime("%Y-%m-%d"), "Monto": row['Cuota_Mensual'], "Comprobante": url}])
                            conn.update(worksheet="Pagos", data=pd.concat([df_h, new_h], ignore_index=True))
                            df_p.at[idx, "Pagos_Realizados"] += 1; df_p.at[idx, "Saldo_Restante"] = round(row["Saldo_Restante"] - row["Cuota_Mensual"], 2)
                            if df_p.at[idx, "Pagos_Realizados"] >= row["Meses_Totales"]: df_p.at[idx, "Estado"] = "PAGADO"
                            conn.update(worksheet="Prestamos", data=df_p)
                            if row.get('Email'): enviar_mail(row['Email'], row['Nombre'], generar_excel_personal(df_p.loc[idx], pd.concat([df_h, new_h]), "PRÉSTAMO"), url, "Prestamos")
                            st.session_state.pago_key += 1; st.rerun()
            with c2: 
                st.download_button(f"📊 EXCEL {row['Nombre'].split()[0]}", data=generar_excel_personal(row, df_h, "PRÉSTAMO"), file_name=f"Reporte_{row['Nombre']}.xlsx", key=f"dlp_{row['ID']}")
            with st.popover("🗑️ ELIMINAR"):
                if st.button("CONFIRMAR BORRADO", key=f"del_p_{row['ID']}"):
                    df_p = df_p[df_p["ID"] != row["ID"]]; conn.update(worksheet="Prestamos", data=df_p); st.rerun()

# --- MODO COOPERATIVA ---
elif sec == "🤝 COOPERATIVA":
    st.title("🤝 COOPERATIVA")
    df_s, df_ph = cargar("Cooperativa"), cargar("Pagos_Coop")
    v_x = st.number_input("💵 VALOR CUOTA FIJA:", value=10.0)
    if not df_s.empty: st.download_button("📊 EXCEL GRUPAL", data=generar_excel_grupal(df_s, "COOPERATIVA"), file_name="Reporte_SanBlas_Coop.xlsx", use_container_width=True)
    
    if st.button("👤 NUEVO SOCIO", key="btn_nuevo_c"): st.session_state.show_form_c = not st.session_state.get('show_form_c', False)
    if st.session_state.get('show_form_c'):
        with st.form("form_c"):
            n, c, e = st.text_input("Nombre:"), st.text_input("Cédula:"), st.text_input("Email:")
            if st.form_submit_button("AÑADIR"):
                new = pd.DataFrame([{"ID":str(uuid.uuid4())[:5], "Nombre":n, "Cedula":c, "Email":e, "Saldo_Total_Aportado":0}])
                conn.update(worksheet="Cooperativa", data=pd.concat([df_s, new], ignore_index=True)); st.session_state.show_form_c = False; st.rerun()

    bq_c = st.text_input("🔍 BUSCAR:")
    act_c = df_s if not df_s.empty else pd.DataFrame()
    if bq_c and not act_c.empty: act_c = act_c[act_c['Nombre'].str.contains(bq_c, case=False)]

    for idx, row in act_c.iterrows():
        with st.expander(f"👤 {row['Nombre'].upper()} | TOTAL: ${row['Saldo_Total_Aportado']}", expanded=(st.session_state.id_abierto == row['ID'])):
            c1, c2 = st.columns(2)
            with c1:
                with st.form(key=f"fc_{row['ID']}"):
                    m = st.number_input("Monto:", value=v_x); ft = st.file_uploader("📸 RECIBO:", key=f"ic_{row['ID']}_{st.session_state.pago_key}")
                    if st.form_submit_button("✅ GUARDAR"):
                        if ft:
                            url = subir_img(ft.getvalue())
                            st.session_state.id_abierto = row['ID']
                            new_h = pd.DataFrame([{"ID_Socio": row['ID'], "Fecha": datetime.now().strftime("%Y-%m-%d"), "Monto": m, "Comprobante": url}])
                            conn.update(worksheet="Pagos_Coop", data=pd.concat([df_ph, new_h], ignore_index=True))
                            df_s.at[idx, "Saldo_Total_Aportado"] = float(row['Saldo_Total_Aportado']) + m
                            conn.update(worksheet="Cooperativa", data=df_s)
                            if row.get('Email'): enviar_mail(row['Email'], row['Nombre'], generar_excel_personal(df_s.loc[idx], pd.concat([df_ph, new_h]), "COOPERATIVA"), url, "Cooperativa")
                            st.session_state.pago_key += 1; st.rerun()
            with c2: st.download_button(f"📊 EXCEL", data=generar_excel_personal(row, df_ph, "COOPERATIVA"), file_name=f"Reporte_{row['Nombre']}.xlsx", key=f"dlc_{row['ID']}")
            with st.popover("🗑️ BORRAR"):
                if st.button("CONFIRMAR BORRADO", key=f"del_c_{row['ID']}"):
                    df_s = df_s[df_s["ID"] != row["ID"]]; conn.update(worksheet="Cooperativa", data=df_s); st.rerun()

# --- MODO AYUDAS ECONÓMICAS ---
elif sec == "🚑 AYUDAS ECON.":
    st.title("🚑 AYUDAS ECONÓMICAS")
    df_a, df_ah = cargar("Ayudas_Listado"), cargar("Pagos_Ayudas")
    v_y = st.number_input("💵 VALOR APORTE:", value=5.0)
    
    col_t1, col_t2 = st.columns([2, 1])
    with col_t1:
        if not df_a.empty: st.download_button("📊 EXCEL GRUPAL", data=generar_excel_grupal(df_a, "AYUDAS ECONÓMICAS"), file_name="Reporte_SanBlas_Ayudas.xlsx", use_container_width=True)
    with col_t2:
        if st.button("🔴 GASTO CAJA", type="secondary"): st.session_state.eg_ay = not st.session_state.get('eg_ay', False)

    if st.session_state.get('eg_ay'):
        with st.form("eg_ay"):
            det_e = st.text_input("Detalle:"); mon_e = st.number_input("Monto:", min_value=1.0)
            if st.form_submit_button("⚠️ RETIRAR"): st.success(f"Gasto guardado."); st.session_state.eg_ay = False; time.sleep(1); st.rerun()

    if st.button("👤 NUEVO COMPAÑERO", key="btn_nuevo_a"): st.session_state.show_form_a = not st.session_state.get('show_form_a', False)
    if st.session_state.get('show_form_a'):
        with st.form("na"):
            n, c, e = st.text_input("Nombre:"), st.text_input("Cédula:"), st.text_input("Email:")
            if st.form_submit_button("AÑADIR"):
                new = pd.DataFrame([{"ID":str(uuid.uuid4())[:5], "Nombre":n, "Cedula":c, "Email":e, "Saldo_Total_Aportado":0}])
                conn.update(worksheet="Ayudas_Listado", data=pd.concat([df_a, new], ignore_index=True)); st.session_state.show_form_a = False; st.rerun()

    bq_a = st.text_input("🔍 BUSCAR:")
    act_a = df_a if not df_a.empty else pd.DataFrame()
    if bq_a and not act_a.empty: act_a = act_a[act_a['Nombre'].str.contains(bq_a, case=False)]

    for idx, row in act_a.iterrows():
        with st.expander(f"👤 {row['Nombre'].upper()} | TOTAL: ${row['Saldo_Total_Aportado']}", expanded=(st.session_state.id_abierto == row['ID'])):
            c1, c2 = st.columns(2)
            with c1:
                with st.form(key=f"fa_{row['ID']}"):
                    m = st.number_input("Monto:", value=v_y); ft = st.file_uploader("📸 RECIBO:", key=f"ia_{row['ID']}_{st.session_state.pago_key}")
                    if st.form_submit_button("✅ GUARDAR"):
                        if ft:
                            url = subir_img(ft.getvalue())
                            st.session_state.id_abierto = row['ID']
                            new_h = pd.DataFrame([{"ID_Socio": row['ID'], "Fecha": datetime.now().strftime("%Y-%m-%d"), "Monto": m, "Comprobante": url}])
                            conn.update(worksheet="Pagos_Ayudas", data=pd.concat([df_ah, new_h], ignore_index=True))
                            df_a.at[idx, "Saldo_Total_Aportado"] = float(row['Saldo_Total_Aportado']) + m
                            conn.update(worksheet="Ayudas_Listado", data=df_a)
                            if row.get('Email'): enviar_mail(row['Email'], row['Nombre'], generar_excel_personal(df_a.loc[idx], pd.concat([df_ah, new_h]), "AYUDAS"), url, "Ayudas")
                            st.session_state.pago_key += 1; st.rerun()
            with c2: st.download_button(f"📊 EXCEL", data=generar_excel_personal(row, df_ah, "AYUDAS"), file_name=f"Reporte_{row['Nombre']}.xlsx", key=f"dla_{row['ID']}")
            with st.popover("🗑️ BORRAR"):
                if st.button("CONFIRMAR BORRADO", key=f"del_a_{row['ID']}"):
                    df_a = df_a[df_a["ID"] != row["ID"]]; conn.update(worksheet="Ayudas_Listado", data=df_a); st.rerun()
